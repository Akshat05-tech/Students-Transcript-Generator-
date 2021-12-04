import os,re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook 
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
import streamlit as st
import smtplib,ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

def marksheet_rollwise(master_roll,answer_responses,correct,wrong):
  s="ANSWER"
  l=list(answer_responses["Roll Number"])
  if s not in l:
    st.warning("No roll number with ANSWER is present, Cannot Process!")
  else:
    st.write("Generating marksheets...")
    x=l.index("ANSWER")
    correct_ans=list(answer_responses.iloc[x,7:])
    path="marksheets"
    if not os.path.isdir(path):
      os.mkdir(path)
    actual_score={}
    google_score={}
    statusAns={}
    for r in range(len(master_roll)):
      roll=master_roll.iloc[r,0]
      name=master_roll.iloc[r,1]
      student_ans=[]
      if(roll in l):
        i=l.index(roll)
        student_ans=list(answer_responses.iloc[i,7:])
      else:
        student_ans=[""]*len(correct_ans)
        #print(roll)
      wb=Workbook()
      ws=wb.worksheets[0]
      img=openpyxl.drawing.image.Image("iitplogo.png")
      ws.add_image(img)
      ws['c7']="MarkSheet"
      ws['c7'].font=Font(bold=True,underline='single',size=12)
      #d=list(answer_responses.iloc[i,[3,6]])
      ft1=Font(bold=True,size=12)
      ft2=Font(color="00FF0000",size=10,bold=True)
      ft3=Font(color="0000FF00",size=10,bold=True)
      ft4=Font(color="000000FF",size=10,bold=True)
      ws['A9']="Name:"
      ws['B9']=name
      ws['B9'].font=ft1
      ws['D9']="Exam:"
      ws['E9']="Quiz"
      ws['E9'].font=ft1
      ws['A10']="Roll Number:"
      ws['B10']=roll
      if(roll not in l):
        ws['f9']="Absent"
        ws["f9"].font=ft2
      na=0        #no of not attempted questions
      crt=0       #no of correct questions
      wrng=0      #no of wrong questions
      ws['B8'].font=ft1
      for c,s in zip(correct_ans,student_ans):
        if(s!=s or s==""):
          na+=1
        elif(c==s):
          crt+=1
        else:
          wrng+=1
      ws['B12']="Right"
      ws['c12']="Wrong"
      ws['d12']="Not Attempt"
      ws['e12']='Max'
      ws['b12'].font=ws['c12'].font=ws['d12'].font=ws['e12'].font=ft1
      ws['a13']='No.'
      ws['a14']='Marking'
      ws['a15']='Total'
      ws['a13'].font=ws['a14'].font=ws['a15'].font=ft1
      ws['b13']=crt
      ws['b14']=correct
      ws['b15']=round(correct*crt,2)
      ws['b13'].font=ws['b14'].font=ws['b15'].font=ft3
      ws['c13']=wrng
      ws['c14']=wrong
      ws['c15']=round(wrong*wrng,2)
      ws['c13'].font=ws['c14'].font=ws['c15'].font=ft2
      ws['d13']=na
      ws['d14']=0
      ws['e13']=len(correct_ans)
      ws['e15']=round(correct*crt+wrong*wrng,2)
      ws['e15'].font=ft4
      max_correct=round(len(correct_ans)*correct,3)
      ws['a18']="Student Ans"
      ws['b18']="Correct Ans"
      ws['a18'].font=ws['b18'].font=ft1
      idx=19
      google_score[roll]=str(round(correct*crt,2))+"/"+str(max_correct)  
      actual_score[roll]=str(round(correct*crt+wrong*wrng,2))+"/"+str(max_correct)
      statusAns[roll]="["+str(crt)+","+str(wrng)+","+str(na)+"]"
      for c,s in zip(correct_ans,student_ans):
        ws["b"+str(idx)]=c
        ws["b"+str(idx)].font=ft4
        if(s!=s):
          continue
        if(c==s):
          ws["a"+str(idx)]=s
          ws["a"+str(idx)].font=ft3
        else:
          ws["a"+str(idx)]=s
          ws["a"+str(idx)].font=ft2
        idx+=1

      out_path=path+"/"+roll.upper()+".xlsx"
      wb.save(out_path)
      wb.close()
    st.write("All marksheets successfully generated!!")

#Preparing concise marksheet.csv
def concise_marksheet(master_roll,answer_responses,correct,wrong):
  s="ANSWER"
  l=list(answer_responses["Roll Number"])
  if s not in l:
    st.warning("No roll number with ANSWER is present, Cannot Process!")
  else:
    x=l.index("ANSWER")
    correct_ans=list(answer_responses.iloc[x,7:])
    path="marksheets"
    if not os.path.isdir(path):
      os.mkdir(path)
    actual_score={}
    google_score={}
    statusAns={}
    for r in range(len(master_roll)):
      roll=master_roll.iloc[r,0]
      name=master_roll.iloc[r,1]
      student_ans=[]
      if(roll in l):
        i=l.index(roll)
        student_ans=list(answer_responses.iloc[i,7:])
      else:
        student_ans=[""]*len(correct_ans)
        #print(roll)

      na=0        #no of not attempted questions
      crt=0       #no of correct questions
      wrng=0      #no of wrong questions
      for c,s in zip(correct_ans,student_ans):
        if(s!=s or s==""):
          na+=1
        elif(c==s):
          crt+=1
        else:
          wrng+=1
      max_correct=round(len(correct_ans)*correct,3)
      idx=19
      google_score[roll]=str(round(correct*crt,2))+"/"+str(max_correct)  
      actual_score[roll]=str(round(correct*crt+wrong*wrng,2))+"/"+str(max_correct)
      statusAns[roll]="["+str(crt)+","+str(wrng)+","+str(na)+"]"
  act_score=[]
  statAns=[]
  g_score=[]
  concise=answer_responses.copy()
  concise.drop(concise.columns[[2]],axis=1,inplace=True)
  for i in range(len(answer_responses)):
    roll=answer_responses.iloc[i,6]
    act_score.append(actual_score[roll])
    statAns.append(statusAns[roll])
    g_score.append(google_score[roll])
  concise.insert(2,"Google_Score",g_score)
  concise.insert(6,"Score_After_Negative",act_score)
  concise.insert(36,"statusAns",statAns)
  idx=len(concise)
  for i in range(len(master_roll)):
    roll=master_roll.iloc[i,0]
    name=master_roll.iloc[i,1]
    if(roll not in l):
      concise.loc[idx]=["","","",name,"","",actual_score[roll],roll]+[""]*len(correct_ans)+[statusAns[roll]]
      idx+=1
  concise.to_csv(path+"/"+"concise_marksheet.csv")


#Dictionary of roll number to email ids => {"1901EE09":["Name","Email id1","Email id2"]}
def emails(answer_responses,sender_name,sender_email,password):
  #regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
  #if re.fullmatch(regex,sender_email):
  st.write("Sending emails...")
  roll_email={}
  #roll_email["1901CB25"]=["Aarooj Yashin","aaroojyashin@gmail.com","md_1901cb25@iitp.ac.in"]
  for i in range(len(answer_responses)):
    d=answer_responses.iloc[i,[1,3,4,6]]
    roll_email[d[3].upper()]=[d[1],d[0],d[2]]
  path="marksheets"
  smtp = smtplib.SMTP('smtp.gmail.com', 587)
  context = ssl.create_default_context()
  smtp.starttls(context=context)
  smtp.login(sender_email,password)
  for i in roll_email.keys():
    for j in range(2):
      msg = MIMEMultipart()
      msg['From'] = sender_email
      msg['To'] = roll_email[i][j+1]
      msg['Date'] = formatdate(localtime = True)
      msg['Subject'] = "Latest Quiz Marksheet"
      text=f"Dear {roll_email[i][0]}, \n Please find your attached marksheet for the latest quiz!! \n Regards,\n {sender_name}"
      msg.attach(MIMEText(text))
      '''if(i=="1901CB25"): 
        file_path=path+"/"+"1401ME61.xlsx"
      else:'''
      file_name=i+".xlsx"
      file_path=path+"/"+file_name
      part = MIMEBase('application', 'vnd.ms-excel')
      part.set_payload(open(file_path,"rb").read())
      encoders.encode_base64(part)
      part.add_header('Content-Disposition', f'attachment; filename = {file_name}')
      msg.attach(part)

      #context = ssl.SSLContext(ssl.PROTOCOL_SSLv3)
      try:
        # We log in into our Google account
        smtp.sendmail(sender_email,roll_email[i][j+1], msg.as_string())
        #print('Email sent!')
      except Exception as e:
        st.write(f'Oh no! Something bad happened!\n{e}')
        break
  smtp.quit()
  st.write("Mails sent successfully!!")